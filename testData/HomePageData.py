import openpyxl
class HomePageData:
    test_HomePage_data = [{"firstname": "Jane", "lastname": "Vo", "email": "a@b.com", "gender": "Female"}, {"firstname": "Hue", "lastname": "Minh", "email": "1@b.com", "gender": "Male"}]

    @staticmethod
    def get_test_data(test_case_name): #self is not required for static methods
        Dict = {}
        wb = openpyxl.load_workbook("../testData/PythonDemo.xlsx")  # "..\\testData\\PythonDemo.xlsx" Either slash will work.
        sheet = wb['Sheet1']
        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column+1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        #print(Dict)
        return [Dict]
