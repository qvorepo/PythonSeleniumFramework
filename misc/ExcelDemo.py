import openpyxl

class ExcelDemo:
    def __init__(self):
        self.wb = openpyxl.load_workbook("../testData/PythonDemo.xlsx") #"..\\testData\\PythonDemo.xlsx" Either slash will work.
        self.sheet = self.wb['Sheet1']
        cell = self.sheet.cell(row=1, column=2)
        print(cell.value + " " + (self.sheet.cell(row=2, column=2)).value)
        print("{}{}".format("Number of rows: ", self.sheet.max_row))
        print("{}{}".format("Numbers of columns: ", self.sheet.max_column))

    def print_sheet_data(self):
        for i in range(1, self.sheet.max_row+1):
            for j in range(1, self.sheet.max_column+1):
                print((self.sheet.cell(row=i, column=j)).value) #print((self.sheet.cell(row=i, column=1)).value)

excelDemo = ExcelDemo()
excelDemo.print_sheet_data()